"""
Excel CRM-report parser.

Reads the first worksheet of an .xlsx export from the CRM, walks the rows
with forward-fill on `Client name` and `Mobile` (CRM exports often leave
those blank on continuation rows that belong to the previous client), and
returns a flat list of normalized event dicts ready for ingestion.

Kept openpyxl-only — no pandas — because the deployment is single-worker
and we don't want to pay the ~80 MB pandas import on every cold start.
"""
import logging
from datetime import datetime, date, time
from typing import Optional

from openpyxl import load_workbook

from app.crm_logic import (
    normalize_mobile,
    normalize_sales_name,
    normalize_stage,
    match_sales_user,
)

log = logging.getLogger(__name__)


# Header → canonical key. Lookup is case-insensitive after trim. Aliases
# cover what CRMs in this market actually export. Anything not in here
# is ignored on the assumption it's a column we don't read.
_HEADER_ALIASES = {
    "client_name": {"client name", "client", "name", "customer", "customer name", "full name", "fullname"},
    "mobile":      {"mobile", "phone", "mobile number", "phone number", "tel", "telephone"},
    "stage":       {"stage", "status"},
    "follow_date": {"follow date", "date", "follow up date", "followup date", "follow-up date", "follow up"},
    "sales_rep":   {"sales rep", "sales", "agent", "representative", "sales agent", "owner"},
    "comment":     {"comment", "comments", "notes", "note", "remarks"},
}

# These four are absolutely required — without them the upload is rejected
# upfront so we never write partial data and pretend it processed cleanly.
_REQUIRED_COLUMNS = ("client_name", "mobile", "stage", "follow_date", "sales_rep")
# Comment is treated as optional — CRMs sometimes ship without it.

# Common string formats the CRM might write when Excel doesn't auto-coerce
# to a real datetime. Order matters: most specific first.
_DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
    "%d-%m-%Y %H:%M:%S",
    "%d-%m-%Y %H:%M",
    "%d-%m-%Y",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
)


def _cell_text(v) -> str:
    """Trim a cell value to a string, handling None and openpyxl quirks."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _parse_follow_date(v) -> Optional[datetime]:
    """Accept a real datetime/date from Excel, or a string in common formats."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        # Bare date → midnight. Keeps downstream code simple (everything is
        # a datetime) without lying about a time the sheet didn't actually
        # carry.
        return datetime.combine(v, time.min)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _resolve_headers(header_row) -> dict:
    """Map each canonical key to its column index in the sheet (1-based).

    Returns {canonical_key: col_index}. Missing required columns raise.
    """
    found: dict = {}
    for idx, cell in enumerate(header_row, start=1):
        if cell is None:
            continue
        label = str(cell).strip().lower()
        if not label:
            continue
        for canon, aliases in _HEADER_ALIASES.items():
            if label in aliases and canon not in found:
                found[canon] = idx
                break

    missing = [k for k in _REQUIRED_COLUMNS if k not in found]
    if missing:
        raise ValueError(
            "Required column(s) missing from sheet header: "
            + ", ".join(missing)
            + ". Headers found: "
            + ", ".join(_cell_text(c) for c in header_row if c is not None)
        )
    return found


def _find_sheet_and_header(wb):
    """Return (worksheet, header_row_tuple, rows_iterator_after_header, first_data_row_num).

    Strategy:
    1. Prefer a sheet whose name contains 'feedback' (case-insensitive).
    2. Try each sheet in order, scanning every row to find the first one
       that contains ALL required canonical columns.  This handles files
       where empty rows or a merged title row precede the real header.
    """
    required_canons = set(_REQUIRED_COLUMNS)

    def _canons_in_row(row):
        found = set()
        for cell in row:
            if cell is None:
                continue
            label = str(cell).strip().lower()
            for canon, aliases in _HEADER_ALIASES.items():
                if label in aliases:
                    found.add(canon)
                    break
        return found

    # Sort sheets: "feedback"-named sheets first, others in workbook order.
    sheets = sorted(
        wb.worksheets,
        key=lambda s: 0 if "feedback" in s.title.lower() else 1,
    )

    for ws in sheets:
        all_rows = list(ws.iter_rows(values_only=True))
        for row_idx, row in enumerate(all_rows):
            if required_canons.issubset(_canons_in_row(row)):
                # This row has ALL required columns — it's the header.
                return ws, row, iter(all_rows[row_idx + 1:]), row_idx + 2

    raise ValueError(
        "No valid header row found in any sheet. "
        "Expected columns: Client name, Mobile, Stage, Follow Date, Sales Rep."
    )


def parse_crm_excel(file_stream, campaign_id: int, conn) -> dict:
    """Parse an .xlsx CRM export into normalized event dicts.

    Args
    ----
    file_stream
        Anything openpyxl.load_workbook accepts (a path or a BytesIO). The
        blueprint passes a BytesIO seeded from request.files['file'].read()
        so we never hit disk.
    campaign_id
        Used for the per-campaign mapping lookup in normalize_stage and
        match_sales_user.
    conn
        Live DB connection — used to read stage_mappings and
        sales_rep_mappings inside the normalization helpers.

    Returns
    -------
    {
      "rows": [ {row_number, client_name, mobile, raw_stage, normalized_stage,
                 follow_date, raw_sales_rep_name, sales_user_id, comment} ],
      "warnings": [str, ...],
      "unmatched_sales_reps": [str, ...],   # de-duplicated, original casing
      "unmatched_stages": [str, ...],       # de-duplicated, original casing
      "total_rows_in_sheet": int,
    }

    The forward-fill behavior is the key bit — see the loop below for the
    last_client_name / last_mobile pattern. CRM exports routinely leave
    those blank on "continuation rows" that belong to the previous client.
    """
    # data_only=True so formula cells (rare in CRM exports, but possible)
    # come through as the cached value instead of "=SUM(...)".
    # read_only=False so _find_sheet_and_header can materialise all rows as a
    # list (read_only streaming iterators can't be rewound).
    wb = load_workbook(file_stream, read_only=False, data_only=True)
    if not wb.worksheets:
        raise ValueError("Workbook has no worksheets.")

    ws, header_row, rows_iter, first_data_row_number = _find_sheet_and_header(wb)

    headers = _resolve_headers(list(header_row))

    def _val(row_tuple, canon):
        idx = headers.get(canon)
        if idx is None or idx > len(row_tuple):
            return None
        return row_tuple[idx - 1]

    out_rows: list = []
    warnings: list = []
    unmatched_reps: dict = {}    # norm → display
    unmatched_stages: dict = {}  # norm → display
    total_rows = 0

    last_client_name: Optional[str] = None
    last_mobile_normalized: Optional[str] = None

    for row_index, row in enumerate(rows_iter, start=first_data_row_number):
        if row is None:
            continue
        total_rows += 1

        raw_client = _cell_text(_val(row, "client_name"))
        raw_mobile = _val(row, "mobile")
        raw_stage = _cell_text(_val(row, "stage"))
        raw_follow = _val(row, "follow_date")
        raw_rep = _cell_text(_val(row, "sales_rep"))
        raw_comment = _cell_text(_val(row, "comment"))

        # Forward-fill: a fresh client header row writes the buffer; a
        # continuation row reads it. We deliberately update the buffer
        # BEFORE checking for "totally blank row" — a row with only a
        # client name and nothing else is a legitimate header pass.
        if raw_client:
            last_client_name = raw_client
        client_name = last_client_name

        # Normalize current row's mobile if present; otherwise inherit.
        if raw_mobile not in (None, ""):
            this_mobile_norm = normalize_mobile(raw_mobile)
            if this_mobile_norm:
                last_mobile_normalized = this_mobile_norm
            else:
                warnings.append(
                    f"Row {row_index}: mobile {raw_mobile!r} couldn't be normalized; "
                    "row will be skipped if no prior valid mobile is in scope."
                )
        mobile = last_mobile_normalized

        # A row with no event-shaped content is just whitespace — skip
        # silently. A row that DOES have content but no mobile in scope
        # is dropped with a warning (we can't attach the event to a lead).
        has_event = bool(raw_stage or raw_follow or raw_rep or raw_comment)
        if not has_event:
            continue

        if not mobile:
            warnings.append(
                f"Row {row_index}: event row has no mobile in scope (header row "
                "before it didn't have a normalizable mobile either). Skipping."
            )
            continue

        # Stage and rep can both fail to resolve; the row is still ingested
        # with NULLs for the unresolved field and the raw value preserved.
        norm_stage = normalize_stage(raw_stage, campaign_id=campaign_id, conn=conn) if raw_stage else None
        if raw_stage and not norm_stage:
            unmatched_stages.setdefault(raw_stage.lower(), raw_stage)

        sales_user_id = match_sales_user(raw_rep, campaign_id, conn) if raw_rep else None
        if raw_rep and sales_user_id is None:
            # Use the normalized form as the dedup key so "Mahmoud  Amr "
            # and "mahmoud amr" don't both show up in the warnings list.
            unmatched_reps.setdefault(normalize_sales_name(raw_rep), raw_rep)

        follow_date = _parse_follow_date(raw_follow)
        if raw_follow not in (None, "") and follow_date is None:
            warnings.append(
                f"Row {row_index}: follow_date {raw_follow!r} couldn't be parsed; stored as NULL."
            )

        out_rows.append({
            "row_number": row_index,
            "client_name": client_name or "",
            "mobile": mobile,
            "raw_stage": raw_stage or None,
            "normalized_stage": norm_stage,
            "follow_date": follow_date,
            "raw_sales_rep_name": raw_rep or None,
            "sales_user_id": sales_user_id,
            "comment": raw_comment or None,
        })

    return {
        "rows": out_rows,
        "warnings": warnings,
        "unmatched_sales_reps": list(unmatched_reps.values()),
        "unmatched_stages": list(unmatched_stages.values()),
        "total_rows_in_sheet": total_rows,
    }
